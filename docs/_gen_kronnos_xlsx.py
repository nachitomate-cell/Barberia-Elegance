"""
Genera docs/kronnos-configuracion.xlsx — template para el dueño de Kronnos.
Hojas: Instrucciones · Servicios x3 · Funcionarios · Premios · Descuentos · Rangos
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"C:\Users\56983\OneDrive\Desktop\Barberia-Elegance\docs\kronnos-configuracion.xlsx"

# ─────────────────────────────────────────────────────────────────
# COLORES POR SEDE (consistente con la app)
# ─────────────────────────────────────────────────────────────────
C_PENA   = "E11D2A"  # rojo Peñablanca
C_LIMA   = "F97316"  # naranja Limache
C_WOMAN  = "EC4899"  # magenta Woman
C_DARK   = "0F172A"  # negro corporativo
C_ACCENT = "FFC700"  # dorado para callouts
C_WHITE  = "FFFFFF"
C_LIGHT  = "F8FAFC"  # gris para filas alternas
C_BORDER = "E2E8F0"

# ─────────────────────────────────────────────────────────────────
# Helpers de estilo
# ─────────────────────────────────────────────────────────────────
def header_style(cell, bg_hex):
    cell.font = Font(bold=True, color=C_WHITE, size=11, name="Segoe UI")
    cell.fill = PatternFill("solid", fgColor=bg_hex)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = Border(
        bottom=Side(style="medium", color=C_DARK),
    )

def title_cell(ws, row, text, bg_hex, span=6):
    ws.cell(row=row, column=1, value=text)
    cell = ws.cell(row=row, column=1)
    cell.font = Font(bold=True, color=C_WHITE, size=18, name="Segoe UI")
    cell.fill = PatternFill("solid", fgColor=bg_hex)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 36

def subtitle_cell(ws, row, text, span=6):
    ws.cell(row=row, column=1, value=text)
    cell = ws.cell(row=row, column=1)
    cell.font = Font(italic=True, color="64748B", size=10, name="Segoe UI")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 22

def row_fill_alt(ws, row, n_cols, fill_hex=C_LIGHT):
    for c in range(1, n_cols + 1):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=fill_hex)

def thin_border():
    return Border(
        left=Side(style="thin", color=C_BORDER),
        right=Side(style="thin", color=C_BORDER),
        top=Side(style="thin", color=C_BORDER),
        bottom=Side(style="thin", color=C_BORDER),
    )

def add_borders(ws, start_row, end_row, n_cols):
    b = thin_border()
    for r in range(start_row, end_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = b

def widths(ws, ws_widths):
    for col, w in ws_widths.items():
        ws.column_dimensions[col].width = w

def example_row(ws, row, values, italic=True):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(italic=italic, color="64748B", size=10, name="Segoe UI")
        c.alignment = Alignment(vertical="center", indent=1)

def data_row(ws, row, values):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(size=10, name="Segoe UI")
        c.alignment = Alignment(vertical="center", indent=1)

# ─────────────────────────────────────────────────────────────────
# Workbook
# ─────────────────────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════════════════════════════════════
# HOJA 1 — Instrucciones
# ════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "📋 Instrucciones"
ws.sheet_properties.tabColor = C_DARK
widths(ws, {"A": 4, "B": 90})

title_cell(ws, 1, "  KRONNOS · Configuración del sistema SynapTech", C_DARK, span=2)
subtitle_cell(ws, 2, "  Template para llenar antes de salir a producción · v1.0 · Junio 2026", span=2)

intro = [
    ("",""),
    ("","Hola! Este archivo tiene todo lo que necesito de tu parte para terminar"),
    ("","la configuración de Kronnos en SynapTech. Lo armé separado por hojas"),
    ("","para que puedas avanzar por partes — no es necesario llenarlo todo de una."),
    ("",""),
    ("","HOJAS:"),
    ("","  1.  Servicios Peñablanca   → catálogo de servicios de la sede Peñablanca"),
    ("","  2.  Servicios Limache       → catálogo de servicios de la sede Limache"),
    ("","  3.  Servicios Woman          → catálogo completo de la sede Woman"),
    ("","  4.  Funcionarios            → nombre + correo de cada profesional"),
    ("","  5.  Premios                  → programa de fidelidad (Club Kronnos)"),
    ("","  6.  Descuentos & Promos     → ofertas fijas o estacionales"),
    ("","  7.  Rangos Silver/Gold/Plat → beneficios por nivel de cliente"),
    ("","  8.  Datos generales         → Instagram, Google reviews, teléfonos"),
    ("",""),
    ("","CÓMO LLENAR:"),
    ("","  • Las filas en GRIS son ejemplos. Déjalas o reemplázalas, da igual."),
    ("","  • Las columnas con encabezado de color son las que debes completar."),
    ("","  • Si no tienes algún dato, déjalo en blanco — lo gestionamos después."),
    ("","  • Cuando termines, me devuelves el archivo por WhatsApp o correo."),
    ("",""),
    ("","¿DUDAS?"),
    ("","  Contáctame: +56 9 8356 8212 · ignaciiio.mate@gmail.com"),
    ("","  SynapTech SpA · synaptechspa.cl · @synaptechspa"),
]

for i, (a, b) in enumerate(intro, start=3):
    cb = ws.cell(row=i, column=2, value=b)
    cb.font = Font(size=11, name="Segoe UI", color="334155")
    cb.alignment = Alignment(vertical="center")
    ws.row_dimensions[i].height = 18

# Banner final
final_row = 3 + len(intro) + 1
ws.cell(row=final_row, column=2, value="🚀  Cuando tengas dudas o termines, escríbeme.")
final_cell = ws.cell(row=final_row, column=2)
final_cell.font = Font(bold=True, size=12, name="Segoe UI", color=C_WHITE)
final_cell.fill = PatternFill("solid", fgColor=C_DARK)
final_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[final_row].height = 36

ws.sheet_view.showGridLines = False

# ════════════════════════════════════════════════════════════════
# Función helper para hojas de servicios
# ════════════════════════════════════════════════════════════════
def hoja_servicios(wb, nombre, color_hex, sede_nombre, categorias, ejemplos_existentes, faltantes_aprox):
    ws = wb.create_sheet(nombre)
    ws.sheet_properties.tabColor = color_hex
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 38, "B": 22, "C": 13, "D": 15, "E": 50})

    title_cell(ws, 1, f"  Servicios · {sede_nombre}", color_hex, span=5)
    subtitle_cell(ws, 2, f"  Catálogo completo de servicios · faltan ~{faltantes_aprox} por cargar", span=5)

    # Headers
    headers = ["Nombre del servicio", "Categoría", "Duración (min)", "Precio (CLP)", "Descripción (opcional)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        header_style(c, color_hex)
    ws.row_dimensions[4].height = 30

    # Validación de categorías (dropdown)
    dv = DataValidation(type="list", formula1=f'"{",".join(categorias)}"', allow_blank=True)
    dv.error = "Categoría no válida"
    dv.errorTitle = "Selecciona una categoría"
    ws.add_data_validation(dv)
    dv.add(f"B5:B100")

    # Ejemplos existentes
    row = 5
    for ej in ejemplos_existentes:
        example_row(ws, row, ej)
        row_fill_alt(ws, row, 5)
        ws.cell(row=row, column=1).comment = None
        row += 1

    # Marca "← ya cargados" en la última fila de ejemplos
    if ejemplos_existentes:
        marker = ws.cell(row=4, column=6, value=f"↑ {len(ejemplos_existentes)} ya cargados")
        marker.font = Font(italic=True, size=9, color="64748B", name="Segoe UI")

    # Filas en blanco para llenar (40 filas)
    for r in range(row, row + 40):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(size=10, name="Segoe UI")
            cell.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r].height = 18

    add_borders(ws, 4, row + 40 - 1, 5)

    # Formato de precio
    for r in range(5, row + 40):
        ws.cell(row=r, column=4).number_format = '"$"#,##0'

# ════════════════════════════════════════════════════════════════
# HOJA 2-4 — Servicios por sede
# ════════════════════════════════════════════════════════════════
categorias_barberia = ["PACKS KRONNOS", "Servicios Masculinos", "Servicios Femeninos", "Manicure", "Otro"]
categorias_woman    = ["Manicura", "Masajes", "Maquillaje", "Pestañas", "Cabello", "Otro"]

ejemplos_pena = [
    ("Pack Toallas Calientes", "PACKS KRONNOS",      45, 24990, "Pack premium con ritual de toallas"),
    ("Corte y Barba",          "Servicios Masculinos", 40, 18990, "Corte + arreglo de barba"),
    ("Corte Masculino",        "Servicios Masculinos", 30, 12990, "Corte clásico"),
]
ejemplos_lima = [
    ("Pack Toallas Calientes",   "PACKS KRONNOS",      45, 24990, ""),
    ("Corte Masculino",          "Servicios Masculinos", 30, 12990, ""),
    ("Corte y Barba",            "Servicios Masculinos", 40, 18990, ""),
    ("Precisión Masculino",      "Servicios Masculinos", 35, 15990, ""),
    ("Corte Bebé",               "Servicios Masculinos", 20,  8990, "Niños menores de 4 años"),
    ("Corte Escolar",            "Servicios Masculinos", 25, 10990, ""),
]
ejemplos_woman = [
    ("Maquillaje Noche",     "Maquillaje", 60, 35000, ""),
    ("Masaje Corporal",      "Masajes",    60, 32000, "Masaje descontracturante"),
    ("Manicura Rusa",        "Manicura",   90, 28000, "Técnica E-File"),
]

hoja_servicios(wb, "1️⃣ Servicios Peñablanca", C_PENA, "Kronnos Studio Peñablanca",
               categorias_barberia, ejemplos_pena, faltantes_aprox=20)
hoja_servicios(wb, "2️⃣ Servicios Limache", C_LIMA, "Kronnos Studio Limache",
               categorias_barberia, ejemplos_lima, faltantes_aprox=6)
hoja_servicios(wb, "3️⃣ Servicios Woman", C_WOMAN, "Kronnos Woman",
               categorias_woman, ejemplos_woman, faltantes_aprox=83)

# ════════════════════════════════════════════════════════════════
# HOJA 5 — Funcionarios
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet("4️⃣ Funcionarios")
ws.sheet_properties.tabColor = C_DARK
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 24, "C": 35, "D": 20, "E": 22})

title_cell(ws, 1, "  Funcionarios · Nombre + correo + teléfono", C_DARK, span=5)
subtitle_cell(ws, 2, "  Cada profesional recibirá notificaciones de citas y podrá tener su acceso al sistema", span=5)

headers = ["Sede", "Nombre del profesional", "Correo electrónico", "Teléfono (opcional)", "Rol"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=i, value=h)
    header_style(c, C_DARK)
ws.row_dimensions[4].height = 30

# Dropdowns
dv_sede = DataValidation(type="list",
                          formula1='"Kronnos Peñablanca,Kronnos Limache,Kronnos Woman"',
                          allow_blank=True)
ws.add_data_validation(dv_sede)
dv_sede.add("A5:A50")

dv_rol = DataValidation(type="list",
                         formula1='"Barbero,Estilista,Manicurista,Esteticista,Maquilladora,Otro"',
                         allow_blank=True)
ws.add_data_validation(dv_rol)
dv_rol.add("E5:E50")

# Pre-poblar con nombres conocidos
preload = [
    ("Kronnos Peñablanca", "Martin",           "", "", "Barbero"),
    ("Kronnos Peñablanca", "Evelyn Contreras", "", "", "Estilista"),
    ("Kronnos Peñablanca", "Araceli",          "", "", "Estilista"),
    ("Kronnos Limache",    "Evelyn Contreras", "", "", "Estilista"),
    ("Kronnos Limache",    "Claudio",          "", "", "Barbero"),
    ("Kronnos Limache",    "Cristian Orostica","", "", "Barbero"),
    ("Kronnos Limache",    "Orlando Palacios", "", "", "Barbero"),
    ("Kronnos Limache",    "Víctor",           "", "", "Barbero"),
    ("Kronnos Woman",      "Kelly",            "", "", "Manicurista"),
    ("Kronnos Woman",      "Ernesto",          "", "", "Otro"),
    ("Kronnos Woman",      "Heydee",           "", "", "Estilista"),
]
for i, row_data in enumerate(preload):
    r = 5 + i
    data_row(ws, r, row_data)
    if i % 2 == 0:
        row_fill_alt(ws, r, 5)

# Filas en blanco extra
for r in range(5 + len(preload), 5 + len(preload) + 20):
    for c in range(1, 6):
        ws.cell(row=r, column=c).font = Font(size=10, name="Segoe UI")
    ws.row_dimensions[r].height = 18

add_borders(ws, 4, 5 + len(preload) + 19, 5)

# ════════════════════════════════════════════════════════════════
# HOJA 6 — Premios (Club Kronnos)
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet("5️⃣ Premios")
ws.sheet_properties.tabColor = C_ACCENT
ws.sheet_view.showGridLines = False
widths(ws, {"A": 18, "B": 34, "C": 16, "D": 50})

title_cell(ws, 1, "  Premios · Club Kronnos", C_DARK, span=4)
subtitle_cell(ws, 2, "  Programa de fidelidad: cuántos sellos acumular y qué se canjea con ellos", span=4)

headers = ["Sede", "Nombre del premio", "Sellos requeridos", "Descripción / qué incluye"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=i, value=h)
    header_style(c, C_DARK)
ws.row_dimensions[4].height = 30

dv_sede2 = DataValidation(type="list",
                           formula1='"Todas,Kronnos Peñablanca,Kronnos Limache,Kronnos Woman"',
                           allow_blank=True)
ws.add_data_validation(dv_sede2)
dv_sede2.add("A5:A40")

# Ejemplos pre-cargados (cursivos para que se vea que son sugerencia)
ejemplos_premios = [
    ("Todas", "Descuento 30%",           5,  "30% de descuento en cualquier servicio. Válido por 30 días."),
    ("Todas", "Servicio gratis",         10, "Servicio gratis (hasta el valor del corte clásico)."),
    ("Todas", "Pack premium",            15, "Pack de toallas + corte + barba con valor diferenciado."),
]
for i, ej in enumerate(ejemplos_premios):
    r = 5 + i
    example_row(ws, r, ej)
    row_fill_alt(ws, r, 4)

# Filas en blanco
for r in range(5 + len(ejemplos_premios), 5 + len(ejemplos_premios) + 15):
    for c in range(1, 5):
        ws.cell(row=r, column=c).font = Font(size=10, name="Segoe UI")
    ws.row_dimensions[r].height = 18

add_borders(ws, 4, 5 + len(ejemplos_premios) + 14, 4)

# ════════════════════════════════════════════════════════════════
# HOJA 7 — Descuentos y Promos
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet("6️⃣ Descuentos & Promos")
ws.sheet_properties.tabColor = C_ACCENT
ws.sheet_view.showGridLines = False
widths(ws, {"A": 18, "B": 32, "C": 28, "D": 14, "E": 40})

title_cell(ws, 1, "  Descuentos & Promociones", C_DARK, span=5)
subtitle_cell(ws, 2, "  Descuentos automáticos o promociones por temporada/cliente", span=5)

headers = ["Sede", "Nombre del descuento", "Cuándo aplica", "% o monto", "Detalle"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=i, value=h)
    header_style(c, C_DARK)
ws.row_dimensions[4].height = 30

dv_sede3 = DataValidation(type="list",
                           formula1='"Todas,Kronnos Peñablanca,Kronnos Limache,Kronnos Woman"',
                           allow_blank=True)
ws.add_data_validation(dv_sede3)
dv_sede3.add("A5:A30")

# Ejemplos pre-cargados
ejemplos_desc = [
    ("Todas", "Cumpleaños",          "Día del cumpleaños del cliente", "10%", "Aplicable a cualquier servicio"),
    ("Todas", "Cliente nuevo",       "Primer servicio en Kronnos",     "15%", "Sin cupo, válido en todas las sedes"),
    ("Todas", "Lunes promocional",   "Todos los lunes",                "2x1", "En cortes masculinos básicos"),
]
for i, ej in enumerate(ejemplos_desc):
    r = 5 + i
    example_row(ws, r, ej)
    row_fill_alt(ws, r, 5)

for r in range(5 + len(ejemplos_desc), 5 + len(ejemplos_desc) + 15):
    for c in range(1, 6):
        ws.cell(row=r, column=c).font = Font(size=10, name="Segoe UI")
    ws.row_dimensions[r].height = 18

add_borders(ws, 4, 5 + len(ejemplos_desc) + 14, 5)

# ════════════════════════════════════════════════════════════════
# HOJA 8 — Rangos
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet("7️⃣ Rangos Silver-Gold-Platinum")
ws.sheet_properties.tabColor = C_ACCENT
ws.sheet_view.showGridLines = False
widths(ws, {"A": 16, "B": 24, "C": 60})

title_cell(ws, 1, "  Rangos · Beneficios por nivel de cliente", C_DARK, span=3)
subtitle_cell(ws, 2, "  El sistema sube de rango automáticamente según servicios acumulados", span=3)

headers = ["Rango", "Servicios para alcanzar", "Beneficios automáticos"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=i, value=h)
    header_style(c, C_DARK)
ws.row_dimensions[4].height = 30

# Pre-cargado con propuesta estándar
rangos_default = [
    ("🥈 Silver",   "0 servicios",  "Beneficio base · acceso al Club Kronnos."),
    ("🥇 Gold",     "10 servicios", "Doble sello en cada servicio. Sugerido: +1 beneficio extra."),
    ("💎 Platinum", "25 servicios", "Doble sello + 10% descuento permanente en todos los servicios."),
]
for i, ej in enumerate(rangos_default):
    r = 5 + i
    data_row(ws, r, ej)
    ws.cell(row=r, column=1).font = Font(bold=True, size=11, name="Segoe UI")
    ws.row_dimensions[r].height = 24
    if i % 2 == 0:
        row_fill_alt(ws, r, 3)

# Nota
nota_row = 5 + len(rangos_default) + 1
ws.cell(row=nota_row, column=1, value="  💡 Los servicios para alcanzar Gold/Platinum son sugerencia. Puedes ajustarlos.")
nota = ws.cell(row=nota_row, column=1)
nota.font = Font(italic=True, size=10, color="92400E", name="Segoe UI")
nota.fill = PatternFill("solid", fgColor="FEF3C7")
ws.merge_cells(start_row=nota_row, start_column=1, end_row=nota_row, end_column=3)
ws.row_dimensions[nota_row].height = 28

add_borders(ws, 4, 5 + len(rangos_default) - 1, 3)

# ════════════════════════════════════════════════════════════════
# HOJA 9 — Datos generales (Instagram, Google reviews, teléfonos)
# ════════════════════════════════════════════════════════════════
ws = wb.create_sheet("8️⃣ Datos generales")
ws.sheet_properties.tabColor = C_DARK
ws.sheet_view.showGridLines = False
widths(ws, {"A": 22, "B": 22, "C": 50})

title_cell(ws, 1, "  Datos generales por sede", C_DARK, span=3)
subtitle_cell(ws, 2, "  Instagram, Google reviews y teléfono de cada sede", span=3)

headers = ["Sede", "Campo", "Valor"]
for i, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=i, value=h)
    header_style(c, C_DARK)
ws.row_dimensions[4].height = 30

filas_grales = [
    ("Kronnos Peñablanca", "Instagram (usuario)",           ""),
    ("Kronnos Peñablanca", "Google reviews URL",             ""),
    ("Kronnos Peñablanca", "Teléfono (ya tengo +56982504870)", "✓ Ya está cargado"),
    ("Kronnos Limache",    "Instagram (usuario)",            ""),
    ("Kronnos Limache",    "Google reviews URL",              ""),
    ("Kronnos Limache",    "Teléfono (ya tengo +56920241041)", "✓ Ya está cargado"),
    ("Kronnos Woman",      "Instagram (usuario)",            ""),
    ("Kronnos Woman",      "Google reviews URL",              ""),
    ("Kronnos Woman",      "Teléfono (FALTA)",                ""),
]
for i, ej in enumerate(filas_grales):
    r = 5 + i
    data_row(ws, r, ej)
    if i % 2 == 0:
        row_fill_alt(ws, r, 3)
    ws.row_dimensions[r].height = 22

# Notita amigable
nota_row = 5 + len(filas_grales) + 1
notas = [
    "🔗 ¿Cómo conseguir el link de Google reviews?",
    "   1. Busca tu sede en Google Maps.",
    "   2. Toca el botón 'Escribir reseña'.",
    "   3. Copia el link de tu navegador y pégalo en la celda.",
    "",
    "📷 ¿Cuál Instagram pongo si las 3 sedes comparten cuenta?",
    "   Pones la misma para las 3. Sin problema.",
]
for i, line in enumerate(notas):
    cell = ws.cell(row=nota_row + i, column=1, value=line)
    cell.font = Font(italic=True, size=10, color="475569", name="Segoe UI")
    ws.merge_cells(start_row=nota_row + i, start_column=1, end_row=nota_row + i, end_column=3)
    ws.row_dimensions[nota_row + i].height = 18

add_borders(ws, 4, 5 + len(filas_grales) - 1, 3)

# ─────────────────────────────────────────────────────────────────
# Guardar
# ─────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Generado: {OUT}")
